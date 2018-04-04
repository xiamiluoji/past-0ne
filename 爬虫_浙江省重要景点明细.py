# python 3 浙江杭州景点信息爬取得，并分类。包括景点名称	、景点描述、	详细描述连接、	景点详细信息	、分类、	等级、	地址
import requests ,os,re
import openpyxl 
import logging
logging.basicConfig(level = logging.DEBUG,format=' %(asctime)s - %(levelname)s -%(message)s')
logging.disable(logging.CRITICAL)
logging.debug('start of program')
logging.basicConfig(filename='myprogramlog.txt',level = logging.DEBUG,format=' %(asctime)s - %(levelname)s -%(message)s')

wb = openpyxl.load_workbook('E:\\杭州\\0314\\浙江省重要景点明细.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

def getaddress(str):
    if str.find('位于') > 0:
        num = str.find('位于')
        str_list = str[num : num+ 70]
    elif str.find('地处') >0:
        num = str.find('地处')
        str_list = str[num : num+ 70]
    elif str.find('地址') > 0:
        num = str.find('地址')
        str_list = str[num : num+ 70]
    else:str_list = "######"
    return str_list
def getAAA(str):
    regex = re.compile(r'AAA.*?级')
    #mm = regex.find(str)
    if regex.findall(str) is not None:
        mm = '_'.join(regex.findall(str))
        return mm
    else: return '无等级'
    

for r in range(sheet.max_row-1):
    str1 = sheet.cell(row =  r + 2,column = 4).value
    sheet.cell(row = r + 2,column = 7).value = getaddress(str1)
    logging.debug(getaddress(str1))
    sheet.cell(row = r + 2,column = 6).value = getAAA(str1)
    logging.debug(getAAA(str1))
wb.save('E:\\杭州\\0314\\浙江省重要景点明细_qing2.xlsx')