
# coding: utf-8

# In[ ]:


import requests ,os,bs4
import openpyxl
#山西教师招聘信息爬取
t = 0 
wb = openpyxl.load_workbook('E:\\杭州\\数据库\\mmm.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = '内容'
sheet['B1'] = '时间'
sheet['C1'] = '连接'
for page in range(1,61):
    if page == 0 :
        res = requests.get('http://sx.offcn.com/html/jiaoshi/zhaokaoxinxi/index.html')
    else:
        res = requests.get('http://sx.offcn.com/html/jiaoshi/zhaokaoxinxi/' + str(page)+ '.html')
    #+ str(page)+
    res.encoding = 'GBK'
    soup = bs4.BeautifulSoup(res.text)
    # 新建excel表格
    
    
    num = len(soup.select('.zg_list a'))
    #print(len(soup.select('#tcjs')))
    #将页面中的信息填写到表格中
    add = t
    for r in range(num):
        sheet.cell(row = r + add + 2, column = 1).value = soup.select('.zg_list a')[r].text
        sheet.cell(row = r + add + 2 , column = 2).value = soup.select('.zg_list span')[r].text
        sheet.cell(row = r + add + 2 , column = 3).value = soup.select('.zg_list a')[r]['href']
        t = t + 1

wb.save('E:\\杭州\\数据库\\mmm29.xlsx')

